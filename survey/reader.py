import csv
import zipfile
import xml.etree.ElementTree as ET


class SurveyReaderMixin:
    def extract_hwpx_text(self, file_path):
        lines = []

        with zipfile.ZipFile(file_path, "r") as archive:
            names = [
                name
                for name in archive.namelist()
                if (
                    name.lower().startswith("contents/section")
                    and name.lower().endswith(".xml")
                )
            ]

            if not names:
                names = [
                    name
                    for name in archive.namelist()
                    if name.lower().endswith(".xml")
                ]

            for name in sorted(names):
                data = archive.read(name)
                root = ET.fromstring(data)

                for element in root.iter():
                    local_name = self.xml_local_name(element.tag)

                    if local_name in {"t", "text"} and element.text:
                        value = element.text.strip()
                        if value:
                            lines.append(value)

        return "\n".join(lines)

    def extract_hwp_text(self, file_path):
        try:
            import win32com.client
        except ImportError as error:
            raise RuntimeError(
                "HWP 읽기를 위해 pywin32가 필요합니다."
            ) from error

        hwp = None

        try:
            hwp = win32com.client.DispatchEx(
                "HWPFrame.HwpObject"
            )
            hwp.XHwpWindows.Item(0).Visible = False

            # 한글 자동화 보안 모듈 등록
            # 설치 환경에 따라 모듈명이 달라 두 이름을 순차 시도
            security_registered = False
            for module_name in [
                "FilePathCheckerModuleExample",
                "FilePathCheckerModule",
            ]:
                try:
                    if hwp.RegisterModule(
                        "FilePathCheckDLL",
                        module_name,
                    ):
                        security_registered = True
                        break
                except Exception:
                    pass

            # 파일 열기 경고창을 억제하는 보조 설정
            try:
                hwp.SetMessageBoxMode(0x00010000)
            except Exception:
                pass

            if hwp.Open(str(file_path)) is False:
                raise RuntimeError("HWP 파일을 열지 못했습니다.")

            text = ""

            # 표 셀 구분이 비교적 잘 남는 형식을 우선 시도
            for format_name in ["UNICODE", "TEXT"]:
                try:
                    text = hwp.GetTextFile(format_name, "")
                except Exception:
                    text = ""

                if text and text.strip():
                    break

            if not security_registered:
                self.status_label.setText(
                    "한글 보안모듈이 등록되지 않아 보안확인창이 표시될 수 있습니다."
                )

            return text or ""

        finally:
            if hwp is not None:
                try:
                    hwp.Quit()
                except Exception:
                    pass

    def extract_spreadsheet_text(self, file_path):
        suffix = file_path.suffix.lower()
        lines = []

        if suffix == ".csv":
            with file_path.open(
                "r",
                encoding="utf-8-sig",
                newline="",
            ) as stream:
                for row in csv.reader(stream):
                    lines.append(" | ".join(str(cell) for cell in row))
            return "\n".join(lines)

        from openpyxl import load_workbook

        workbook = load_workbook(
            file_path,
            data_only=True,
            read_only=True,
        )
        sheet = workbook[workbook.sheetnames[0]]

        for row in sheet.iter_rows(values_only=True):
            values = [
                "" if value is None else str(value)
                for value in row
            ]
            lines.append(" | ".join(values))

        workbook.close()
        return "\n".join(lines)

    def read_survey_file_text(self, file_path):
        suffix = file_path.suffix.lower()

        if suffix == ".hwp":
            return self.extract_hwp_text(file_path)

        if suffix == ".hwpx":
            return self.extract_hwpx_text(file_path)

        if suffix in {".xlsx", ".xlsm", ".csv"}:
            return self.extract_spreadsheet_text(file_path)

        if suffix == ".txt":
            return file_path.read_text(
                encoding="utf-8-sig"
            )

        raise RuntimeError(
            f"지원하지 않는 파일 형식입니다: {suffix}"
        )
